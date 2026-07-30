"""
Agent 6: PDF下载 + 研究包生成
下载Top-N论文PDF，生成完整研究包
"""
import os
import re
import asyncio
from datetime import datetime


async def download_pdfs(browser_mgr, papers: list[dict], output_dir: str,
                        max_count: int = 20) -> list[dict]:
    """
    下载评分最高的论文PDF。
    仅在用户已有合法访问权限时自动下载。
    """
    pdf_dir = os.path.join(output_dir, "PDF")
    os.makedirs(pdf_dir, exist_ok=True)

    downloaded = []
    to_download = papers[:max_count]

    print(f"\n  [Agent6] 准备下载 {len(to_download)} 篇PDF...")

    for i, paper in enumerate(to_download):
        title = paper.get("title", f"paper_{i+1}")
        url = paper.get("url", "")
        safe_title = re.sub(r'[\\/:*?"<>|]', '_', title)[:60]
        filename = f"{i+1:02d}_{safe_title}.pdf"
        filepath = os.path.join(pdf_dir, filename)

        if os.path.exists(filepath):
            print(f"    [{i+1}/{len(to_download)}] 已存在: {filename}")
            downloaded.append({"title": title, "file": filepath, "status": "exists"})
            continue

        if not url:
            downloaded.append({"title": title, "file": "", "status": "no_url"})
            continue

        print(f"    [{i+1}/{len(to_download)}] 下载: {title[:40]}...")

        try:
            page = await browser_mgr.new_page()
            try:
                from paper import ensure_detail_page
                await ensure_detail_page(browser_mgr.session, page, url)
                await browser_mgr.session.dismiss_known_overlays(page)

                # 检查是否有下载权限
                not_logged = await page.locator('.downloadlink.icon-notlogged, [class*="notlogged"]').count()
                if not_logged > 0:
                    print(f"      [SKIP] 无下载权限（需要登录/机构授权）")
                    downloaded.append({"title": title, "file": "", "status": "no_permission"})
                    continue

                # 尝试点击PDF下载
                pdf_link = page.locator('#pdfDown, .btn-dlpdf a').first
                if await pdf_link.count() > 0:
                    async with page.expect_download(timeout=30000) as download_info:
                        await pdf_link.click()
                    download = await download_info.value
                    await download.save_as(filepath)
                    print(f"      [OK] 已保存: {filename}")
                    downloaded.append({"title": title, "file": filepath, "status": "ok"})
                else:
                    print(f"      [SKIP] 未找到PDF下载按钮")
                    downloaded.append({"title": title, "file": "", "status": "no_button"})

                await asyncio.sleep(2)
            finally:
                await page.close()

        except Exception as e:
            err = str(e)[:60]
            print(f"      [FAIL] {err}")
            downloaded.append({"title": title, "file": "", "status": f"error: {err}"})

    ok_count = sum(1 for d in downloaded if d["status"] in ("ok", "exists"))
    print(f"  [Agent6] 下载完成: {ok_count}/{len(to_download)} 成功")
    return downloaded


def generate_research_package(output_dir: str, topic: str, request, strategy,
                               all_papers: list, scored_papers: list,
                               author_analysis, downloads: list):
    """生成完整研究包：Excel + Markdown报告"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    os.makedirs(output_dir, exist_ok=True)

    # === 1. 论文信息.xlsx ===
    _save_papers_excel(scored_papers, os.path.join(output_dir, "论文信息.xlsx"))

    # === 2. 作者分析.xlsx ===
    _save_author_excel(author_analysis, os.path.join(output_dir, "作者分析.xlsx"))

    # === 3. 参考文献.xlsx (全部检索结果) ===
    _save_references_excel(all_papers, os.path.join(output_dir, "参考文献.xlsx"))

    # === 4. 研究报告.md ===
    _save_report_md(output_dir, topic, request, strategy, all_papers,
                    scored_papers, author_analysis, downloads)

    print(f"\n  [Agent6] 研究包已生成: {output_dir}")
    print(f"    ├── PDF/ ({sum(1 for d in downloads if d['status'] in ('ok','exists'))} 篇)")
    print(f"    ├── 论文信息.xlsx")
    print(f"    ├── 作者分析.xlsx")
    print(f"    ├── 参考文献.xlsx")
    print(f"    └── 研究报告.md")


def _save_papers_excel(scored_papers, filepath):
    """保存评分后的论文列表"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "推荐论文"

    headers = ["排名", "评分", "标题", "作者", "单位", "期刊", "年份", "引用", "关键词", "摘要"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for i, sp in enumerate(scored_papers, 1):
        d = sp.data
        authors = d.get("authors", []) or d.get("detailAuthors", [])
        author_str = ", ".join(a.get("name", "") if isinstance(a, dict) else str(a) for a in authors[:5])
        affs = d.get("affiliations", [])
        aff_str = "; ".join(affs[:2]) if affs else ""
        keywords = d.get("keywords", [])
        kw_str = "; ".join(keywords[:5]) if keywords else ""
        date = d.get("date", "") or d.get("pubInfo", "")

        row_data = [i, sp.score, sp.title, author_str, aff_str,
                    d.get("journal", "") or d.get("journalDetail", ""),
                    date, d.get("citations", ""), kw_str, d.get("abstract", "")]

        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i+1, column=j, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 6
    ws.column_dimensions["I"].width = 30
    ws.column_dimensions["J"].width = 60

    wb.save(filepath)


def _save_author_excel(analysis, filepath):
    """保存作者分析"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()

    # Sheet1: 核心作者
    ws = wb.active
    ws.title = "核心作者"
    headers = ["作者", "单位", "论文数", "总引用", "推荐度", "主要期刊", "研究方向"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)

    for i, a in enumerate(analysis.authors[:20], 2):
        ws.cell(row=i, column=1, value=a.name)
        ws.cell(row=i, column=2, value="; ".join(a.affiliations[:2]))
        ws.cell(row=i, column=3, value=a.paper_count)
        ws.cell(row=i, column=4, value=a.total_citations)
        ws.cell(row=i, column=5, value="★" * a.recommendation)
        ws.cell(row=i, column=6, value="; ".join(a.journals[:3]))
        ws.cell(row=i, column=7, value="; ".join(a.keywords[:5]))

    # Sheet2: 机构分布
    ws2 = wb.create_sheet("机构分布")
    ws2.cell(row=1, column=1, value="机构").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="论文数").font = Font(bold=True)
    for i, (inst, count) in enumerate(analysis.institutions, 2):
        ws2.cell(row=i, column=1, value=inst)
        ws2.cell(row=i, column=2, value=count)

    # Sheet3: 高频关键词
    ws3 = wb.create_sheet("高频关键词")
    ws3.cell(row=1, column=1, value="关键词").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="频次").font = Font(bold=True)
    for i, (kw, count) in enumerate(analysis.top_keywords, 2):
        ws3.cell(row=i, column=1, value=kw)
        ws3.cell(row=i, column=2, value=count)

    wb.save(filepath)


def _save_references_excel(all_papers, filepath):
    """保存全部检索结果"""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "全部文献"

    headers = ["序号", "标题", "作者", "期刊", "年份", "引用", "数据库"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)

    for i, p in enumerate(all_papers, 2):
        authors = p.get("authors", [])
        if isinstance(authors, list):
            author_str = ", ".join(a.get("name", "") if isinstance(a, dict) else str(a) for a in authors[:3])
        else:
            author_str = str(authors)
        ws.cell(row=i, column=1, value=i-1)
        ws.cell(row=i, column=2, value=p.get("title", ""))
        ws.cell(row=i, column=3, value=author_str)
        ws.cell(row=i, column=4, value=p.get("journal", ""))
        ws.cell(row=i, column=5, value=p.get("date", ""))
        ws.cell(row=i, column=6, value=p.get("citations", ""))
        ws.cell(row=i, column=7, value=p.get("database", ""))

    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    wb.save(filepath)


def _save_report_md(output_dir, topic, request, strategy, all_papers,
                    scored_papers, author_analysis, downloads):
    """生成文献调研报告"""
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    ok_downloads = sum(1 for d in downloads if d["status"] in ("ok", "exists"))

    lines = [
        f"# 文献调研报告",
        f"",
        f"> 生成时间: {now}",
        f"> 研究主题: {topic}",
        f"",
        f"## 检索概况",
        f"",
        f"| 指标 | 数值 |",
        f"|------|------|",
        f"| 检索主题 | {' + '.join(request.topics)} |",
        f"| 时间范围 | {request.start_year or '不限'} - {request.end_year or '不限'} |",
        f"| 来源要求 | {request.source or '不限'} |",
        f"| 共检索 | {len(all_papers)} 篇 |",
        f"| 评分筛选 | {len(scored_papers)} 篇 |",
        f"| 最终下载 | {ok_downloads} 篇 |",
        f"",
        f"## 检索策略",
        f"",
        f"`",
        f"{strategy.search_expression}",
        f"`",
        f"",
        f"## 推荐论文 (Top {min(20, len(scored_papers))})",
        f"",
    ]

    for i, sp in enumerate(scored_papers[:20], 1):
        stars = "★" * min(5, int(sp.score / 20)) + "☆" * (5 - min(5, int(sp.score / 20)))
        d = sp.data
        journal = d.get("journal", "") or d.get("journalDetail", "")
        lines.append(f"### {i}. {sp.title}")
        lines.append(f"")
        lines.append(f"- 评分: {sp.score:.0f} {stars}")
        lines.append(f"- 期刊: {journal}")
        lines.append(f"- 引用: {d.get('citations', '0')}")
        abstract = d.get("abstract", "")
        if abstract:
            lines.append(f"- 摘要: {abstract[:200]}...")
        lines.append(f"")

    # 作者分析
    lines.extend([
        f"## 核心作者",
        f"",
        f"| 作者 | 单位 | 论文数 | 引用 | 推荐 |",
        f"|------|------|--------|------|------|",
    ])
    for a in author_analysis.authors[:10]:
        aff = a.affiliations[0] if a.affiliations else ""
        stars = "★" * a.recommendation
        lines.append(f"| {a.name} | {aff} | {a.paper_count} | {a.total_citations} | {stars} |")

    # 机构
    lines.extend([f"", f"## 主要机构", f""])
    for inst, count in author_analysis.institutions[:8]:
        lines.append(f"- {inst}: {count}篇")

    # 关键词
    lines.extend([f"", f"## 高频关键词", f""])
    kw_str = ", ".join(f"**{kw}**({c})" for kw, c in author_analysis.top_keywords[:15])
    lines.append(kw_str)

    # 年份分布
    if author_analysis.year_distribution:
        lines.extend([f"", f"## 时间分布", f""])
        for year, count in author_analysis.year_distribution:
            bar = "█" * count
            lines.append(f"- {year}: {bar} ({count})")

    # 下载目录
    lines.extend([f"", f"## 下载目录", f""])
    for d in downloads:
        status_icon = "✅" if d["status"] in ("ok", "exists") else "❌"
        lines.append(f"- {status_icon} {d['title'][:50]}")

    lines.extend([f"", f"---", f"*由 CNKI Scholar Copilot 自动生成*", ""])

    filepath = os.path.join(output_dir, "研究报告.md")
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
